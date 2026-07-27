# -*- coding: utf-8 -*-
"""
מיזוג נתונים: Google Contacts (CSV) + Donations Summary (Excel) -> חוברת מאוחדת.

שימוש:
    python3 tools/merge_sources.py --contacts contacts.csv --donations Donations.xlsx --out starter/crm-donors-filled.xlsx

מה נוצר בחוברת:
    תורמים               — כרטיס לכל תורם מגוגל (שם, טלפון, מייל, כתובת) + דרגת קוויטל שהותאמה
    לבדיקה_קוויטל        — שמות קוויטל שלא הותאמו אוטומטית, עם הצעת תורם לאישור ידני
    קובץ_התאמה_חודשיים   — התורמים החודשיים מהאקסל (שם אנגלי) + שתי עמודות עבריות ריקות למילוי
    מזדמנים              — תורמים מזדמנים (Occasional) לפי חודש
"""
import re, csv, argparse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

_ap = argparse.ArgumentParser()
_ap.add_argument("--contacts", required=True, help="ייצוא Google Contacts בפורמט CSV")
_ap.add_argument("--donations", required=True, help="קובץ Donations Summary (xlsx)")
_ap.add_argument("--out", default="starter/crm-donors-filled.xlsx")
_ap.add_argument("--review-json", default=None, help="ייצוא רשימת ההתאמות-לבדיקה כ-JSON (לדף האישור)")
_ap.add_argument("--approvals", default=None, help="קובץ JSON של אישורים ידניים {שם_קוויטל: {choice}}")
_ap.add_argument("--phone-matches", default=None, help="קובץ JSON {שם_קוויטל: טלפון} לחיבור טלפונים שנמצאו לפי צליל")
_ap.add_argument("--manual-merges", default=None, help="קובץ JSON: רשימת קבוצות שמות למיזוג ידני")
_ap.add_argument("--phone-overrides", default=None, help="קובץ JSON {שם: טלפון} לתיקון ידני של טלפון")
_ap.add_argument("--regular-matches", default=None, help="קובץ JSON {שם אנגלי: שם עברי} לאישור חיבור תורם קבוע")
_ap.add_argument("--regular-review-json", default=None, help="ייצוא הקבועים שלא חוברו + מועמדים (לדף האישור)")
_ap.add_argument("--new-regulars", default=None, help="קובץ JSON {שם אנגלי: שם עברי} ליצירת כרטיסי תורם קבוע חדשים")
_ap.add_argument("--purposes", default=None, help="קובץ JSON {שם תורם: עבור מה תרם}")
_ap.add_argument("--parnes-yom", default=None, help="קובץ JSON {שם תורם: {תאריך, יום, חודש, הקדשה, נוסח}}")
_args = _ap.parse_args()

# סדר החודשים העבריים (שנה מתחילה בתשרי)
_HMONTHS = ['תשרי','חשון','כסלו','טבת','שבט','אדר','ניסן','אייר','סיון','תמוז','אב','אלול']
def hmonth_order(m):
    m = (m or '').replace('מרחשון','חשון').replace('סיוון','סיון').replace('אדר א','אדר').replace('אדר ב','אדר').strip()
    return _HMONTHS.index(m)+1 if m in _HMONTHS else 99
CONTACTS = _args.contacts
DON = _args.donations
OUT = _args.out

NIKUD = re.compile(r'[֑-ׇ]')
KV_SUFFIX = re.compile(r'[\-–—]?\s*ק[וו][ו]?[יי]?[ט]ל.*$')  # " - קוויטל" ווריאציות
TORM_SUFFIX = re.compile(r'\s*תורם\s*$')

def norm(s):
    if not s: return ''
    s = NIKUD.sub('', str(s))
    s = s.replace('"','').replace("'",'').replace('`','')
    s = KV_SUFFIX.sub('', s)
    s = TORM_SUFFIX.sub('', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s

FIN = str.maketrans('ךםןףץ','כמנפצ')
def loose(s):
    """מפתח 'רופף' לספיגת וריאציות איות: אותיות סופיות + וו/יי כפולים."""
    s = norm(s).translate(FIN)
    s = s.replace('וו','ו').replace('יי','י').replace('אא','א')
    s = re.sub(r'\s+','',s)
    return s

def skel(s):
    """שלד עיצורי — מסיר אותיות חלשות (תנועות). לזיהוי כפילויות שם משפחה."""
    s = norm(s).translate(FIN)
    s = re.sub(r'[אהועי]','',s)
    return re.sub(r'\s+','',s)

# ----- תעתיק עברית<->אנגלית לזיהוי שמות מאקסל התרומות -----
_H2L = {'א':'','ב':'b','ג':'g','ד':'d','ה':'','ו':'v','ז':'z','ח':'h','ט':'t','י':'',
        'כ':'k','ך':'k','ל':'l','מ':'m','ם':'m','נ':'n','ן':'n','ס':'s','ע':'','פ':'f',
        'ף':'f','צ':'ts','ץ':'ts','ק':'k','ר':'r','ש':'s','ת':'t'}
def _redu(s):
    s = s.lower()
    for a,b in [('sch','s'),('sh','s'),('tz','s'),('ts','s'),('ch','h'),('ck','k'),
                ('ph','f'),('th','t'),('w','v'),('q','k'),('c','k'),('x','ks')]:
        s = s.replace(a,b)
    return re.sub(r'(.)\1+',r'\1', re.sub(r'[aeiou]','',s))
def hskel(h): return _redu(''.join(_H2L.get(c,'') for c in norm(h)))
def eskel(e): return _redu(re.sub(r'[^a-z]','', str(e or '').lower()))
def edist(a,b):
    if abs(len(a)-len(b))>1: return 9
    dp=list(range(len(b)+1))
    for i,ca in enumerate(a,1):
        prev=dp[0]; dp[0]=i
        for j,cb in enumerate(b,1):
            cur=dp[j]; dp[j]=min(dp[j]+1,dp[j-1]+1,prev+(ca!=cb)); prev=cur
    return dp[-1]

# ---------- קריאת אנשי קשר ----------
rows = list(csv.DictReader(open(CONTACTS, encoding='utf-8')))
def labs(r): return set(l.strip() for l in (r.get('Labels') or '').split(' ::: ') if l.strip())

KV = {'קוויטל','קוויטל ישז','קוויטל 101','קוויטל לא פעיל','יששכר זבולון','קוויטל אב פו'}
DON_L = {'תורמים','תורמים מזדמנים','תורמים בארץ עברית'}

def tier_from_labels(L):
    if 'קוויטל ישז' in L or 'יששכר זבולון' in L: return 'יששכר_זבולון'
    if 'קוויטל 101' in L: return 'קוויטל_101'
    if 'קוויטל' in L or 'קוויטל אב פו' in L: return 'קוויטל_כללי'
    if 'קוויטל לא פעיל' in L: return 'קוויטל_כללי'
    return ''

# מפת דרגה לפי שם עברי, מכרטיסי הקוויטל — לפי מפתח מדויק וגם מפתח רופף
tier_map = {}; tier_loose = {}; kv_names = {}
inactive = set()
order = {'יששכר_זבולון':3,'קוויטל_101':2,'קוויטל_כללי':1,'':0}
for r in rows:
    L = labs(r)
    if L & KV:
        raw = (r.get('First Name','')+' '+r.get('Middle Name','')+' '+r.get('Last Name',''))
        nm = norm(raw); lk = loose(raw); t = tier_from_labels(L)
        if nm and order.get(t,0) >= order.get(tier_map.get(nm,''),0): tier_map[nm]=t
        if lk and order.get(t,0) >= order.get(tier_loose.get(lk,''),0): tier_loose[lk]=t
        if lk: kv_names.setdefault(lk, norm(raw))
        if 'קוויטל לא פעיל' in L: inactive.add(nm)

def lookup_tier(nm, lk):
    if nm in tier_map: return tier_map[nm], 'מדויק'
    if lk in tier_loose: return tier_loose[lk], 'רופף'
    return '', ''
matched_kv_keys = set()

# ---------- חילוץ שמות התפילה מכרטיסי הקוויטל ----------
# השמות שמורים בשדה המותאם 'קוויטל' או בהערות
prayer_rows = []
for r in rows:
    L = labs(r)
    if not (L & KV): continue
    donor_nm = norm(r.get('First Name','')+' '+r.get('Middle Name','')+' '+r.get('Last Name',''))
    cf = (r.get('Custom Field 1 - Value') or '').strip()
    note = (r.get('Notes') or '').strip()
    prayer = cf or note
    if not prayer: continue
    prayer = prayer.replace('\r','').strip()
    prayer = re.sub(r',\s*,', ',', prayer)           # פסיקים כפולים
    # דלג על שאריות שדות טכניים (טלפון/מייל) או ערך מספרי בלבד
    if re.search(r'(Phone|E-?mail|- Value|:::)', prayer): continue
    if re.fullmatch(r'[\d\s\-\+()]+', prayer): continue
    prayer_rows.append({'donor':donor_nm, 'prayer':prayer, 'tier':tier_from_labels(L),
                        'tags':';'.join(sorted(L & KV))})

# ---------- בניית רשומות תורמים ----------
donors = []
seen = {}; seen_loose = set()
def phone_join(r):
    ps=[r.get('Phone 1 - Value',''),r.get('Phone 2 - Value','')]
    raw=' / '.join(p.strip() for p in ps if p.strip())
    return raw.replace(' ::: ',' / ')

def _norm_one_phone(x):
    """מנרמל מספר בודד: אמריקאי -> +1XXXXXXXXXX ; ישראלי -> +972 ; אחר נשמר."""
    x=x.strip()
    if not x: return ''
    has_plus=x.startswith('+')
    digits=re.sub(r'\D','',x)
    if digits.startswith('00'): digits=digits[2:]
    if len(digits)<7: return ''      # שארית לא-תקינה (למשל '+1' בודד) — מושמט
    if digits.startswith('972'):                     # ישראל
        return '+972 '+digits[3:].lstrip('0')
    if has_plus and not digits.startswith('1'):      # מדינה אחרת (בלגיה וכו') — נשמר
        return '+'+digits
    if digits.startswith('0') and len(digits)<=10:   # ישראלי מקומי 05x
        return '+972 '+digits[1:]
    if digits.startswith('1') and len(digits)==11:   # ארה"ב עם קידומת
        digits=digits[1:]
    if len(digits)==10:                              # ארה"ב 10 ספרות
        return f'+1 {digits[:3]}-{digits[3:6]}-{digits[6:]}'
    if has_plus: return '+'+digits
    return x

def normalize_phone(p):
    """מנרמל מחרוזת טלפון (יתכן כמה מספרים מופרדים ב-/), ומסיר כפילויות לפי הספרות."""
    if not p: return ''
    out=[]; seen=set()
    for x in re.split(r'\s*/\s*',p):
        z=_norm_one_phone(x)
        if not z: continue
        sig=re.sub(r'\D','',z)        # חתימה = כל הספרות (מספר זהה לא יופיע פעמיים)
        if sig in seen: continue
        seen.add(sig); out.append(z)
    return ' / '.join(out)

for r in rows:
    L = labs(r)
    if not (L & DON_L):   # רק כרטיסים עם תווית תורמים
        continue
    first=(r.get('First Name','')+' '+r.get('Middle Name','')).strip()
    last=r.get('Last Name','').strip()
    nm = norm(first+' '+last); lk = loose(first+' '+last)
    tier, how = lookup_tier(nm, lk)
    if tier and lk in tier_loose: matched_kv_keys.add(lk)
    addr = r.get('Address 1 - Formatted','') or ', '.join(x for x in [r.get('Address 1 - City',''),r.get('Address 1 - Country','')] if x)
    notes = ' | '.join(x for x in [r.get('Notes',''),r.get('Custom Field 1 - Value','')] if x)
    tags = ';'.join(sorted(L & (KV|DON_L)))
    donors.append({
        'first':re.sub(r'\s+',' ',first).strip(), 'last':last,
        'org':r.get('Organization Name',''),
        'phone':phone_join(r), 'email':r.get('E-mail 1 - Value',''),
        'addr':addr.replace('\n',', '), 'tier':tier, 'how':how, 'tags':tags,
        'bday':r.get('Birthday',''), 'notes':notes,
        'n-flag':'לא פעיל' if nm in inactive else '', 'nm':nm,
    })
    seen[nm]=True; seen_loose.add(lk)

# טעינת אישורים ידניים (מדף האישור)
approvals = {}
if _args.approvals:
    import json as _json
    approvals = {norm(k): v for k, v in _json.load(open(_args.approvals, encoding='utf-8')).items()}
donor_by_norm = {}
for d in donors:
    donor_by_norm[norm(d['first'] + ' ' + d['last'])] = d
n_linked = 0; n_newapproved = 0

# חיבורי טלפון שנמצאו לפי צליל
phone_matches = {}
if _args.phone_matches:
    import json as _json2
    phone_matches = {norm(k): v for k, v in _json2.load(open(_args.phone_matches, encoding='utf-8')).items()}

# אינדקס לפי שם-משפחה רופף מכל אנשי הקשר (לא רק תורמים), להצעת התאמות
all_by_surname={}
for r in rows:
    L=labs(r)
    if (L & KV) and not (L & DON_L):  # דלג על כרטיסי-קוויטל עצמם
        continue
    first=(r.get('First Name','')+' '+r.get('Middle Name','')).strip()
    last=r.get('Last Name','').strip()
    if not (first or last): continue
    entry={'first':norm(first),'last':norm(last),
           'phone':phone_join(r),'is_donor':bool(L & DON_L)}
    sur=loose(last or first.split(' ')[-1])
    if sur: all_by_surname.setdefault(sur,[]).append(entry)

def suggest(fullname):
    toks=norm(fullname).split(' ')
    if not toks: return []
    sur=loose(toks[-1]); firstl=loose(toks[0])
    cands=all_by_surname.get(sur,[])
    ranked=sorted(cands,key=lambda d:(0 if loose(d['first']).startswith(firstl[:2]) else 1,
                                      0 if d['is_donor'] else 1))
    return ranked

# כרטיסי קוויטל ללא כרטיס תורם -> רשימת 'לבדיקה' עם הצעת התאמה
review=[]; kv_only=0
for r in rows:
    L=labs(r)
    if (L & KV) and not (L & DON_L):
        raw=(r.get('First Name','')+' '+r.get('Middle Name','')+' '+r.get('Last Name',''))
        nm=norm(raw); lk=loose(raw)
        # החלת אישור ידני אם קיים
        if nm in approvals:
            ch = approvals[nm].get('choice')
            tier_here = tier_from_labels(L)
            if ch == '__new__':
                donors.append({'first':nm,'last':'','org':'','phone':'','email':'','addr':'',
                    'tier':tier_here,'how':'אושר: חדש','tags':';'.join(sorted(L&KV)),
                    'bday':'','notes':'','n-flag':'תורם חדש (אושר ידנית)','nm':nm})
                n_newapproved += 1
            else:
                dm = donor_by_norm.get(norm(ch))
                if dm is not None:
                    if not dm['tier']: dm['tier'] = tier_here
                    dm['how'] = 'אושר ידנית'
                else:
                    # איש קשר קיים שאינו מתויג 'תורמים' — נוסיף אותו כתורם עם הטלפון שאושר
                    donors.append({'first':ch,'last':'','org':'','phone':approvals[nm].get('phone',''),
                        'email':'','addr':'','tier':tier_here,'how':'אושר ידנית','tags':';'.join(sorted(L&KV)),
                        'bday':'','notes':'','n-flag':'חובר מאיש קשר (אושר ידנית)','nm':norm(ch)})
                    donor_by_norm[norm(ch)] = donors[-1]
                n_linked += 1
            seen[nm]=True; seen_loose.add(lk)
            continue
        if nm and nm not in seen and lk not in seen_loose:
            cands=suggest(nm)
            # ללא הצעת התאמה כלל -> זהו תורם מוכר מהקוויטל, נוסיף אותו לרשימה (בלי טלפון עדיין)
            if not cands:
                donors.append({'first':nm,'last':'','org':'','phone':'','email':'','addr':'',
                    'tier':tier_from_labels(L),'how':'מהקוויטל','tags':';'.join(sorted(L&KV)),
                    'bday':'','notes':'','n-flag':'מהקוויטל — אין טלפון עדיין','nm':nm})
                seen[nm]=True; seen_loose.add(lk); kv_only+=1
                continue
            sug=cands[0] if cands else None
            prayer=(r.get('Custom Field 1 - Value') or r.get('Notes') or '').replace('\r','').strip()
            review.append({'kv':nm,'tier':tier_from_labels(L),'prayer':prayer,
                'sug':(sug['first']+' '+sug['last']).strip() if sug else '',
                'sug_phone':sug['phone'] if sug else '',
                'sug_isdonor':('תורם' if sug and sug['is_donor'] else ('איש קשר' if sug else '')),
                'cands':[{'name':(c['first']+' '+c['last']).strip(),'phone':c['phone'],
                          'is_donor':c['is_donor']} for c in cands[:4]],
                'n':len(cands)})
            kv_only+=1

# ---------- קריאת תרומות (Data + Occasional) ----------
wb=openpyxl.load_workbook(DON, data_only=True)
def s(v):
    if v is None: return ''
    if isinstance(v,float): return str(int(v)) if v==int(v) else str(round(v,2))
    return str(v).strip()

CH={'banquest':'בנק_ווסט','authorize':'אותורייז','checks':'צק','check':'צק','ach':'העברה_בנקאית',
    'fidelity':'פידליטי','ojc':'OJC','donors fund':'דונורס_פאנד','website':'אתר','nedarim':'נדרים_פלוס'}
def chan(m):
    m=(m or '').strip(); return CH.get(m.lower(), m)

data=wb['Data']; data_rows=[]
for r in range(2,data.max_row+1):
    typ=s(data.cell(r,1).value); nm=s(data.cell(r,2).value)
    if not nm: continue
    en_name=s(data.cell(r,4).value); en_sur=s(data.cell(r,5).value)
    # שם עסק: כשעמודת Names אינה שם האדם (לא מכילה את שם המשפחה) — זהו שם עסק
    business = nm if (en_sur and en_sur.lower() not in nm.lower()) else ''
    data_rows.append({'type':typ,'names':nm,'method':chan(s(data.cell(r,3).value)),
        'en_name':en_name,'en_sur':en_sur,'business':business,
        'email':s(data.cell(r,6).value),'amount':s(data.cell(r,7).value)})

occ=wb['Occasional']; occ_rows=[]
months=['ינואר','פברואר','מרץ','אפריל','מאי','יוני','יולי','אוגוסט','ספטמבר','אוקטובר','נובמבר','דצמבר']
for r in range(2,occ.max_row+1):
    nm=s(occ.cell(r,1).value)
    if not nm: continue
    vals={months[i]:s(occ.cell(r,3+i).value) for i in range(12)}
    total=sum(float(v) for v in vals.values() if v and v.replace('.','').isdigit())
    occ_rows.append({'nm':nm,'vals':vals,'total':total})

# ---------- כתיבת החוברת ----------
FONT='Arial'
HF=PatternFill('solid',fgColor='1F4E78'); HFONT=Font(name=FONT,bold=True,color='FFFFFF')
YEL=PatternFill('solid',fgColor='FFF2CC')
thin=Side(style='thin',color='D9D9D9'); BD=Border(left=thin,right=thin,top=thin,bottom=thin)
owb=openpyxl.Workbook()

def sheet(title,headers,rows_data,yellow=()):
    ws=owb.create_sheet(title); ws.sheet_view.rightToLeft=True
    for c,h in enumerate(headers,1):
        cell=ws.cell(1,c,h); cell.fill=HF; cell.font=HFONT
        cell.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True); cell.border=BD
        ws.column_dimensions[get_column_letter(c)].width=max(12,min(30,len(h)+6))
    for c in yellow: ws.cell(1,c).fill=YEL
    for i,row in enumerate(rows_data,2):
        for c,v in enumerate(row,1):
            cell=ws.cell(i,c,v); cell.border=BD; cell.font=Font(name=FONT,size=10)
    ws.freeze_panes='A2'; ws.row_dimensions[1].height=28
    return ws

owb.remove(owb.active)

# מפת שם אנגלי מאקסל התרומות (לפי אימייל) — לקישור לכרטיס העברי
email_to_en = {}
for dr in data_rows:
    em = (dr.get('email') or '').strip().lower()
    en = (dr['en_name'] + ' ' + dr['en_sur']).strip()
    if em and en: email_to_en[em] = en

# חיבור טלפונים שנמצאו לפי צליל (למי שאין לו טלפון) + נרמול + פיצול שם + שם אנגלי
n_phone_added = 0
for d in donors:
    d.setdefault('english', '')
    d.setdefault('aliases', [])
    if not d.get('phone') and d.get('nm') in phone_matches and phone_matches[d['nm']]:
        d['phone'] = phone_matches[d['nm']]
        d['how'] = (d.get('how','') + ' +טלפון').strip()
        d['n-flag'] = 'טלפון חובר לפי צליל'
        n_phone_added += 1
    d['phone'] = normalize_phone(d.get('phone',''))
    if not d.get('last') and d.get('first'):
        toks = d['first'].split()
        if len(toks) >= 2:
            d['first'] = ' '.join(toks[:-1]); d['last'] = toks[-1]
    em = (d.get('email') or '').strip().lower()
    if em in email_to_en: d['english'] = email_to_en[em]

# תיקוני טלפון ידניים (שם -> טלפון נכון, מחליף)
phone_overrides = {}
if _args.phone_overrides:
    import json as _j4
    phone_overrides = {norm(k): v for k, v in _j4.load(open(_args.phone_overrides, encoding='utf-8')).items()}
n_phone_fixed = 0
for d in donors:
    for cand in (norm(d['first']+' '+d['last']), norm(d['last']+' '+d['first'])):
        if cand in phone_overrides:
            d['phone'] = normalize_phone(phone_overrides[cand]); n_phone_fixed += 1; break

# ---------- מיזוג כפילויות (שם משפחה זהה בשלד + שם פרטי דומה) ----------
order_t = {'יששכר_זבולון':3,'קוויטל_101':2,'קוויטל_כללי':1,'':0}
from collections import OrderedDict
groups = OrderedDict()
for d in donors:
    key = (skel(d['last']) or skel(d['first']), loose(d['first']))
    groups.setdefault(key, []).append(d)
merged = []; n_merged = 0
for key, grp in groups.items():
    if len(grp) == 1:
        merged.append(grp[0]); continue
    grp.sort(key=lambda x:(0 if x['phone'] else 1, 0 if x['tier'] else 1, 0 if x.get('email') else 1))
    base = grp[0]
    phones = []
    for d in grp:
        for p in re.split(r'\s*/\s*', d.get('phone','')):
            if p and p not in phones: phones.append(p)
    base['phone'] = ' / '.join(phones)
    base['tier'] = max((d['tier'] for d in grp), key=lambda t:order_t.get(t,0))
    for f in ['org','email','addr','bday','english']:
        if not base.get(f):
            for d in grp:
                if d.get(f): base[f] = d[f]; break
    base['tags'] = ';'.join(sorted({t for d in grp for t in (d.get('tags') or '').split(';') if t}))
    base['n-flag'] = 'מוזג מכפילות'
    merged.append(base); n_merged += len(grp) - 1
donors[:] = merged

# ---------- מיזוגים ידניים שסומנו על ידי המשתמש ----------
def keyset(s):
    return frozenset(loose(t) for t in norm(s).split() if loose(t))
manual_groups = []
if _args.manual_merges:
    import json as _j3
    manual_groups = _j3.load(open(_args.manual_merges, encoding='utf-8'))
n_manual = 0
for group in manual_groups:
    wanted = [keyset(x) for x in group if keyset(x)]
    matched = []
    for d in donors:
        dk = keyset(d['first'] + ' ' + d['last'])
        # התאמה אם שם-הקבוצה מוכל בשם התורם (או להיפך) עם חפיפה של 2+ מילים
        if any((w <= dk or dk <= w) and len(w & dk) >= 2 for w in wanted):
            matched.append(d)
    if len(matched) < 2: continue
    matched.sort(key=lambda x:(0 if x['phone'] else 1, 0 if x['tier'] else 1))
    base = matched[0]
    phones = []
    for d in matched:
        for p in re.split(r'\s*/\s*', d.get('phone','')):
            if p and p not in phones: phones.append(p)
    base['phone'] = ' / '.join(phones)
    base['tier'] = max((d['tier'] for d in matched), key=lambda t:order_t.get(t,0))
    for f in ['org','email','addr','bday','english']:
        if not base.get(f):
            for d in matched:
                if d.get(f): base[f] = d[f]; break
    base.setdefault('aliases', [])
    for d in matched:
        base['aliases'].append(d['first'] + ' ' + d['last'])
    base['n-flag'] = 'מוזג ידנית'
    for d in matched[1:]:
        if d in donors: donors.remove(d)
    n_manual += len(matched) - 1

# ---------- העשרת תרומות + קטגוריות ----------
MONTHS_HE = ['ינואר','פברואר','מרץ','אפריל','מאי','יוני','יולי','אוגוסט','ספטמבר','אוקטובר','נובמבר','דצמבר']
# מטרות (עבור מה תרם) — קלט ידני
purposes = {}
if _args.purposes:
    import json as _j8
    purposes = {norm(k): v for k, v in _j8.load(open(_args.purposes, encoding='utf-8')).items()}
# פרנס יום (לילה מסוים בשנה) — קלט ידני
parnes = {}
if _args.parnes_yom:
    import json as _j9
    parnes = {norm(k): v for k, v in _j9.load(open(_args.parnes_yom, encoding='utf-8')).items()}

for d in donors:
    d['ls'] = hskel(d['last']); d['fs'] = hskel(d['first'])
    for k in ('category','dtype','amount','channel','pay_status','last_active','months','purpose'):
        d.setdefault(k, '')
    d.setdefault('py_list', [])
    for cand in (norm(d['first']+' '+d['last']), norm(d['last']+' '+d['first'])):
        if cand in purposes:
            d['purpose'] = purposes[cand]; break
    for cand in (norm(d['first']+' '+d['last']), norm(d['last']+' '+d['first'])):
        if cand in parnes:
            nights = parnes[cand]
            if isinstance(nights, dict): nights = [nights]
            d['py_list'] = nights
            if not d['purpose']:
                d['purpose'] = ('פרנס יום — ' + nights[0].get('תאריך','')) if len(nights)==1 else f'פרנס יום ({len(nights)} לילות)'
            break

# תורמים קבועים (Data, אנגלית) -> חיבור בתעתיק / אישור ידני
regular_matches = {}
if _args.regular_matches:
    import json as _j5
    regular_matches = {norm(k): v for k, v in _j5.load(open(_args.regular_matches, encoding='utf-8')).items()}

# יצירת כרטיסי תורם קבוע חדשים (שהוזנו ידנית)
n_newreg = 0
if _args.new_regulars:
    import json as _j7
    for eng, heb in _j7.load(open(_args.new_regulars, encoding='utf-8')).items():
        toks = norm(heb).split()
        first = ' '.join(toks[:-1]) if len(toks) >= 2 else heb
        last = toks[-1] if len(toks) >= 2 else ''
        ph = normalize_phone(phone_matches.get(norm(heb), ''))
        card = {'first':first,'last':last,'english':eng,'org':'','phone':ph,'email':'','addr':'',
                'tier':'','how':'חדש (הוזן ידנית)','tags':'','bday':'','notes':'','n-flag':'תורם קבוע חדש',
                'nm':norm(heb),'aliases':[],'category':'','dtype':'','amount':'','channel':'',
                'pay_status':'','last_active':'','ls':hskel(last),'fs':hskel(first),
                'purpose':purposes.get(norm(heb), '')}
        donors.append(card)
        regular_matches[norm(eng)] = heb
        n_newreg += 1

donor_by_name = {}
for d in donors:
    donor_by_name[norm(d['first']+' '+d['last'])] = d
    donor_by_name[norm(d['last']+' '+d['first'])] = d

n_reg = 0; unmatched_data = []
dsheet = wb['Data']
for r in range(2, dsheet.max_row+1):
    if not s(dsheet.cell(r,2).value): continue
    sur = s(dsheet.cell(r,5).value); fn = s(dsheet.cell(r,4).value)
    if not sur: continue
    typ = s(dsheet.cell(r,1).value); method = chan(s(dsheet.cell(r,3).value)); amount = s(dsheet.cell(r,7).value)
    statuses = [s(dsheet.cell(r,8+i).value) for i in range(12)]
    es = eskel(sur); efs = eskel(fn)
    cands = [d for d in donors if d['ls'] and (d['ls']==es or edist(d['ls'],es)<=1)]
    best=None; bestsc=99; uniq=True
    ekey = norm(fn+' '+sur)
    if ekey in regular_matches:
        # אישור ידני: '__none__' = לבטל חיבור (לא לשייך), אחרת לשייך לתורם שצוין
        if regular_matches[ekey] != '__none__':
            best = donor_by_name.get(norm(regular_matches[ekey])); bestsc=0; uniq=True
        else:
            best = None; bestsc=99; uniq=False
    else:
        for d in cands:
            sc = edist(d['ls'],es)*2 + (edist(d['fs'],efs) if (efs and d['fs']) else 2)
            if sc < bestsc: bestsc=sc; best=d; uniq=True
            elif sc == bestsc: uniq=False
    if best is not None and bestsc<=3 and uniq:
        paid = sum(1 for x in statuses if x and 'ccep' in x.lower())
        decl = sum(1 for x in statuses if x and x.strip().upper()=='NR')
        last = max([i for i,x in enumerate(statuses) if x], default=-1)
        def _num(v):
            try: return float(str(v).replace(',',''))
            except: return 0.0
        def _fmt(x): return str(int(x)) if x==int(x) else str(round(x,2))
        if best.get('category') == 'קבוע':
            # תרומה נוספת לאותו תורם — סכום מצטבר
            best['amount'] = _fmt(_num(best.get('amount')) + _num(amount))
            best['dtype'] = (best.get('dtype','') + ' + נוסף') if 'נוסף' not in best.get('dtype','') else best['dtype']
        else:
            best['category'] = 'קבוע'
            best['dtype'] = {'ישז':'שותפות יששכר-זבולון','חודשי':'חודשי'}.get(typ, typ)
            best['amount'] = amount; best['channel'] = method
            best['pay_status'] = (f'שולם {paid} חודשים' + (f', {decl} נדחו' if decl else '')) if (paid or decl) else ''
            best['last_active'] = MONTHS_HE[last] if last>=0 else ''
            # מפת חודשים: p=עבר, x=נדחה, -=ריק
            best['months'] = ''.join('p' if (x and 'ccep' in x.lower()) else 'x' if (x and x.strip().upper()=='NR') else '-' for x in statuses)
        if not best.get('english'): best['english'] = (fn+' '+sur).strip()
        n_reg += 1
    else:
        scored = sorted(cands, key=lambda d: edist(d['ls'],es)*2 + (edist(d['fs'],efs) if (efs and d['fs']) else 2))
        unmatched_data.append({'fn':fn,'sur':sur,'typ':typ,'amount':amount,
            'candidates':[{'name':(d['first']+' '+d['last']).strip(),'phone':d['phone'],'tier':d['tier']} for d in scored[:4]]})

# ייצוא הקבועים שלא חוברו + מועמדים, לדף האישור
if _args.regular_review_json:
    import json as _j6
    payload = {'unmatched':[{'id':f'M{i:04d}','english':(u['fn']+' '+u['sur']).strip(),
                             'type':u['typ'],'amount':u['amount'],'candidates':u['candidates']}
                            for i,u in enumerate(unmatched_data,1)],
               'donors':[{'name':(d['first']+' '+d['last']).strip(),'phone':d['phone'],'tier':d['tier']}
                         for d in sorted(donors,key=lambda x:x['last'] or '')]}
    _j6.dump(payload, open(_args.regular_review_json,'w',encoding='utf-8'), ensure_ascii=False)
    print('JSON אישור קבועים:', _args.regular_review_json, '(', len(unmatched_data), ')')

# מזדמנים (עברית) -> חיבור לפי שם + חודש אחרון פעיל
donor_loose = {}
for d in donors:
    donor_loose.setdefault(loose(d['first']+' '+d['last']), d)
    donor_loose.setdefault(loose(d['last']+' '+d['first']), d)
n_occ = 0
for o in occ_rows:
    d = donor_loose.get(loose(o['nm']))
    if d is None: continue
    if not d['category']: d['category'] = 'מזדמן'
    if not d['amount'] and o['total']: d['amount'] = str(int(o['total']))
    active = [i for i,m in enumerate(months) if o['vals'].get(m)]
    if active: d['last_active'] = MONTHS_HE[max(active)]
    n_occ += 1

# תורמים — מיון לפי שם משפחה (א-ב), עמודת שם משפחה לפני שם פרטי
tor_rows=[]
for i,d in enumerate(sorted(donors,key=lambda x:(x['last'] or 'תתת', x['first'])),1):
    tor_rows.append([f'ת-{i:05d}',d['last'],d['first'],d.get('english',''),d['org'],d['phone'],d['email'],d['addr'],
                     d.get('category',''),d.get('dtype',''),d.get('amount',''),d.get('channel',''),d.get('pay_status',''),d.get('last_active',''),
                     d['tier'],d.get('how',''),d['tags'],d['bday'],d['notes'],d['n-flag'],';'.join(d.get('aliases',[])),d.get('months',''),d.get('purpose','')])
sheet('תורמים',['מזהה_תורם','שם_משפחה_עברי','שם_פרטי_עברי','שם_אנגלי','שם_עסק','טלפון','אימייל','כתובת',
    'קטגוריה','סוג_תרומה','סכום','ערוץ_תשלום','סטטוס_תשלום','פעיל_לאחרונה',
    'דרגת_קוויטל','אופן_התאמה','תוויות_גוגל','יום_הולדת','הערות','סטטוס','כינויים','סטטוס_חודשים','עבור_מה'],tor_rows)

# קובץ התאמה (Data החודשיים) — עם עמודות עבריות ריקות
mt_rows=[]
for i,d in enumerate(data_rows,1):
    mt_rows.append([f'M-{i:04d}',(d['en_name']+' '+d['en_sur']).strip() or d['names'],
        d['business'],d['type'],d['method'],d['amount'],d['email'],'',''])
sheet('קובץ_התאמה_חודשיים',['מפתח','שם_אנגלי','שם_עסק','סוג(ישז/חודשי)','אמצעי_תשלום','סכום_חודשי',
    'אימייל','שם פרטי עברי','שם משפחה עברי'],mt_rows,yellow=(8,9))

# מזדמנים (כבר בעברית)
oc_rows=[]
for i,d in enumerate(occ_rows,1):
    paid=';'.join(f'{m}:{v}' for m,v in d['vals'].items() if v)
    oc_rows.append([f'O-{i:04d}',d['nm'],int(d['total']) if d['total'] else '',paid])
sheet('מזדמנים',['מפתח','שם_עברי','סכום_כולל_השנה','פירוט_חודשי'],oc_rows)

# פרנס יום — לוח שנה לפי תאריך עברי (כמה לילות לתורם)
def gen_nusach(date, ded):
    return ('יהי רצון שזכות הלימודים והתפילות הנעשים כאן בכולל חצות בעת רצון הגדול '
            'של חצות הלילה עד הבוקר' + (', ' + date if date else '') + ', יהיו ויעמדו לזכות '
            + (ded or ''))
allnights = []
for d in donors:
    for n in d.get('py_list', []):
        allnights.append((d, n))
allnights.sort(key=lambda t:(hmonth_order(t[1].get('חודש','')), int(t[1].get('יום')) if str(t[1].get('יום','')).isdigit() else 99))
py_rows = []
for d, n in allnights:
    date = n.get('תאריך',''); ded = n.get('הקדשה','')
    nusach = n.get('נוסח','') or gen_nusach(date, ded)
    py_rows.append([date, n.get('יום',''), n.get('חודש',''), hmonth_order(n.get('חודש','')),
                    (d['first']+' '+d['last']).strip(), ded, nusach, d['phone']])
sheet('פרנס_יום',['תאריך','יום','חודש','סדר','שם_התורם','הקדשה','נוסח_תפילה','טלפון'],py_rows)

# שמות_לתפילה — השמות שהאברכים קוראים, לפי דרגה
sp_rows=[]
torder={'יששכר_זבולון':0,'קוויטל_101':1,'קוויטל_כללי':2}
for i,d in enumerate(sorted(prayer_rows,key=lambda x:(torder.get(x['tier'],9),x['donor'])),1):
    sp_rows.append([f'ש-{i:04d}','',d['donor'],d['prayer'],'',d['tier'],d['tags']])
sp_ws=sheet('שמות_לתפילה',['מזהה_שם','מזהה_תורם','שם_התורם','שמות','בקשה','דרגת_תפילה','תוויות_קוויטל'],sp_rows)
sp_ws.column_dimensions['D'].width=70
for rr in range(2,len(sp_rows)+2):
    sp_ws.cell(rr,4).alignment=Alignment(wrap_text=True,vertical='top')

# לבדיקה — שמות קוויטל ללא כרטיס תורם + הצעת התאמה
rv_rows=[]
for i,d in enumerate(sorted(review,key=lambda x:(x['n']==0,x['tier']!='יששכר_זבולון',x['kv'])),1):
    match_hint = 'התאמה יחידה' if d['n']==1 else (f'{d["n"]} אפשרויות' if d['n']>1 else 'לא נמצא — אולי חדש')
    rv_rows.append([f'R-{i:04d}',d['kv'],d['tier'],d['sug'],d['sug_isdonor'],d['sug_phone'],match_hint,''])
sheet('לבדיקה_קוויטל',['מפתח','שם_בקוויטל','דרגה','שם_מוצע','סוג_מוצע','טלפון_מוצע','הערכת_התאמה','אישור(כן/לא)'],
    rv_rows,yellow=(8,))

owb.save(OUT)

# ---------- ייצוא JSON לדף האישור ----------
if _args.review_json:
    import json
    payload=[]
    for i,d in enumerate(sorted(review,key=lambda x:(x['n']!=1,x['tier']!='יששכר_זבולון',x['kv'])),1):
        payload.append({'id':f'R{i:04d}','kvittel':d['kv'],'tier':d['tier'],
            'prayer':d.get('prayer','')[:220],'candidates':d['cands'],'n':d['n']})
    json.dump(payload,open(_args.review_json,'w',encoding='utf-8'),ensure_ascii=False,indent=1)
    print('JSON לאישור:',_args.review_json,'(',len(payload),'שורות )')

# ---------- דוח ----------
from collections import Counter
tc=Counter(d['tier'] for d in donors if d['tier'])
hc=Counter(d.get('how','') for d in donors if d['tier'])
print('OUT:',OUT)
print('כרטיסי תורם מלאים (מגוגל):',len(donors))
print('  עם דרגת קוויטל שהותאמה אוטומטית:',sum(tc.values()),' ',dict(tc))
print('  אופן ההתאמה:',dict(hc))
if approvals:
    print('אישורים ידניים שהוחלו: חוברו לאיש קשר =',n_linked,'| נוספו כחדשים =',n_newapproved)
if phone_matches:
    print('טלפונים שחוברו לפי צליל:',n_phone_added)
print('כפילויות שמוזגו:',n_merged)
if manual_groups:
    print('מיזוגים ידניים:',n_manual)
print(f'תורמים קבועים שחוברו (מ-Data): {n_reg} | לא חוברו: {len(unmatched_data)}')
print(f'מזדמנים שחוברו: {n_occ}')
print('שמות קוויטל שנשארו לבדיקה:',len(review))
one=sum(1 for d in review if d['n']==1); multi=sum(1 for d in review if d['n']>1); none=sum(1 for d in review if d['n']==0)
print('   התאמה יחידה מוצעת:',one,'| כמה אפשרויות:',multi,'| חדש/לא נמצא:',none)
print('תורמים חודשיים (Data) לקובץ ההתאמה:',len(data_rows))
print('מזדמנים (Occasional):',len(occ_rows))
