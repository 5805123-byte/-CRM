#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
הפקת קוויטל להדפסה מתוך חוברת האקסל של המערכת.

דוגמאות שימוש:
    # כל השמות
    python3 tools/generate_kvittel.py

    # רק קוויטל 101
    python3 tools/generate_kvittel.py --דרגה קוויטל_101

    # רשימה מיוחדת לפי תווית (למשל לנסיעה למירון), כולל חד-פעמיים
    python3 tools/generate_kvittel.py --תווית מירון

    # בחירת קובץ קלט/פלט
    python3 tools/generate_kvittel.py --קלט starter/crm-donors.xlsx --פלט kvittel.xlsx
"""
import argparse
import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

FONT = "Arial"


def read_sheet(wb, name):
    """מחזיר רשימת מילונים {כותרת: ערך} לכל שורה בגיליון."""
    if name not in wb.sheetnames:
        return []
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    out = []
    for r in rows[1:]:
        rec = {headers[i]: r[i] for i in range(len(headers))}
        # דילוג על שורות ריקות (העמודה הראשונה = מזהה)
        if not (r and r[0] not in (None, "")):
            continue
        out.append(rec)
    return out


def main():
    p = argparse.ArgumentParser(description="הפקת קוויטל להדפסה")
    p.add_argument("--קלט", dest="inp", default="starter/crm-donors.xlsx")
    p.add_argument("--פלט", dest="out", default=None)
    p.add_argument("--דרגה", dest="level", default=None,
                   help="יששכר_זבולון / קוויטל_101 / קוויטל_כללי / חד_פעמי")
    p.add_argument("--תווית", dest="label", default=None,
                   help="סינון לפי תווית קוויטל (למשל: מירון)")
    args = p.parse_args()

    wb = openpyxl.load_workbook(args.inp, data_only=True)
    donors = {d.get("מזהה_תורם"): d for d in read_sheet(wb, "תורמים") if d.get("מזהה_תורם")}
    names = read_sheet(wb, "שמות_לתפילה")

    # אם אין עדיין גיליון 'שמות_לתפילה' — בונים קוויטל ישירות מרשימת התורמים לפי דרגה
    if not names:
        selected = []
        for d in donors.values():
            level = d.get("דרגת_קוויטל") or ""
            if not level:
                continue
            labels = str(d.get("תוויות_גוגל") or d.get("תוויות_קוויטל") or "")
            if args.level and level != args.level:
                continue
            if args.label and args.label not in labels:
                continue
            full = ((d.get("שם_פרטי_עברי", "") or "") + " " + (d.get("שם_משפחה_עברי", "") or "")).strip()
            selected.append({"שמות": full, "בקשה": "", "דרגה": level, "תורם": full})
        return finish(selected, args)

    selected = []
    for n in names:
        donor = donors.get(n.get("מזהה_תורם"), {})
        level = n.get("דרגת_תפילה") or donor.get("דרגת_קוויטל") or ""
        labels = str(n.get("תוויות_קוויטל") or donor.get("תוויות_קוויטל") or "")
        if args.level and level != args.level:
            continue
        if args.label and args.label not in labels:
            continue
        selected.append({
            "שמות": n.get("שמות") or "",
            "בקשה": n.get("בקשה") or "",
            "דרגה": level,
            "תורם": (donor.get("שם_פרטי_עברי", "") or "") + " " + (donor.get("שם_משפחה_עברי", "") or ""),
        })

    return finish(selected, args)


def finish(selected, args):
    # מיון: לפי דרגה ואז לפי שם
    order = {"יששכר_זבולון": 0, "קוויטל_101": 1, "קוויטל_כללי": 2, "חד_פעמי": 3}
    selected.sort(key=lambda x: (order.get(x["דרגה"], 9), str(x["שמות"])))

    # כותרת הקוויטל
    title = "קוויטל"
    if args.level:
        title += f" — {args.level.replace('_', ' ')}"
    if args.label:
        title += f" — {args.label}"

    out_wb = openpyxl.Workbook()
    ws = out_wb.active
    ws.title = "קוויטל"
    ws.sheet_view.rightToLeft = True
    ws.sheet_view.showGridLines = False

    ws["A1"] = title
    ws["A1"].font = Font(name=FONT, bold=True, size=18, color="1F4E78")
    ws["A2"] = "הודפס: " + datetime.date.today().isoformat() + f"   |   סה\"כ שמות: {len(selected)}"
    ws["A2"].font = Font(name=FONT, italic=True, size=10, color="808080")

    hdr = ["#", "שמות לתפילה", "בקשה", "דרגה"]
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c, h in enumerate(hdr, 1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.font = Font(name=FONT, bold=True, color="FFFFFF")
        cell.fill = openpyxl.styles.PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
    for i, s in enumerate(selected, 1):
        row = 4 + i
        vals = [i, s["שמות"], s["בקשה"], s["דרגה"].replace("_", " ")]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.font = Font(name=FONT, size=13 if c == 2 else 11)
            cell.alignment = Alignment(horizontal="right" if c == 2 else "center", vertical="center")
            cell.border = border
        ws.row_dimensions[row].height = 22

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 16
    # הגדרות הדפסה
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = "4:4"

    out_path = args.out or ("kvittel-" + datetime.date.today().isoformat() + ".xlsx")
    out_wb.save(out_path)
    print(f"נוצר קוויטל עם {len(selected)} שמות: {out_path}")


if __name__ == "__main__":
    main()
